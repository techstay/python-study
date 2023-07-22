# %%
import random
from pathlib import Path

from faker import Faker
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

fake = Faker("zh_CN")

file = Path("~/Desktop/book.xlsx").expanduser().resolve()

if file.exists():
    book = load_workbook(file)
else:
    book = Workbook()

# %%
ws = book.active
ws.title = "学生成绩表"

# 表头
ws["a1"] = "姓名"
ws["b1"] = "语文"
ws["c1"] = "数学"
ws["d1"] = "英语"
ws["e1"] = "物理"
ws["f1"] = "化学"
ws["g1"] = "生物"


def get_random_score():
    mark = random.gauss(75, 12.5)
    if mark < 0:
        mark = 0
    if mark > 100:
        mark = 100
    return int(mark)


# 生成50个人的成绩
N = 50
for row in range(2, 2 + N):
    ws[f"a{row}"] = fake.name()
    ws[f"b{row}"] = get_random_score()
    ws[f"c{row}"] = get_random_score()
    ws[f"d{row}"] = get_random_score()
    ws[f"e{row}"] = get_random_score()
    ws[f"f{row}"] = get_random_score()
    ws[f"g{row}"] = get_random_score()

# 总计、平均数、方差、标准差
ws[f"a{N+3}"] = "总计"
ws[f"a{N+4}"] = "平均数"
ws[f"a{N+5}"] = "方差"
ws[f"a{N+6}"] = "标准差"

column = ["b", "c", "d", "e", "f", "g"]
for col in column:
    ws[f"{col}{N+3}"] = f"=sum({col}2:{col}{N+1})"
    ws[f"{col}{N+4}"] = f"=average({col}2:{col}{N+1})"
    ws[f"{col}{N+5}"] = f"=varp({col}2:{col}{N+1})"
    ws[f"{col}{N+6}"] = f"=stdevp({col}2:{col}{N+1})"

# 前10位学生的成绩图表

names = Reference(ws, min_col=1, min_row=2, max_row=11)
values = Reference(ws, min_col=2, max_col=7, min_row=1, max_row=11)
chart = BarChart()
chart.add_data(values, titles_from_data=True)
chart.title = "前10位学生的成绩图表"
chart.x_axis.title = "姓名"
chart.y_axis.title = "分数"
chart.set_categories(names)
chart.width = 15
chart.height = 10
ws.add_chart(chart, "J1")
book.save(file)

# %%
